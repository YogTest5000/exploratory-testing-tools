from openpyxl import load_workbook, Workbook
import os

excel_url = r'path'

def write_txt_to_excel(txt_file, excel_file=excel_url):
    # Check if the Excel file exists, otherwise create a new one
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # Read lines from the text file
    with open(txt_file, "r", encoding="utf-8") as file:
        lines = file.readlines()

    # Clean lines and write them as a new row in the Excel sheet
    new_row = [line.strip() for line in lines]

    # Append the new row to the Excel sheet
    ws.append(new_row)

    # Save the updated Excel file
    wb.save(excel_file)
    print(f"Data from {txt_file} has been written to {excel_file}")

# Example usage
txt_file_path = "path" 

write_txt_to_excel(txt_file_path)
