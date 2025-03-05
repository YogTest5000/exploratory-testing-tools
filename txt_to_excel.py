import sys
import os
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QVBoxLayout, QWidget, QMessageBox
from openpyxl import Workbook


class TxtToExcelConverter(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("TXT to Excel Converter")
        self.setGeometry(100, 100, 400, 200)

        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)

        self.layout = QVBoxLayout()
        self.centralWidget.setLayout(self.layout)

        self.loadButton = QPushButton("Load TXT File and Convert")
        self.loadButton.clicked.connect(self.loadTxtFile)
        self.layout.addWidget(self.loadButton)

    def loadTxtFile(self):
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(self, "Open TXT File", "", "Text Files (*.txt);;All Files (*)",
                                                  options=options)

        if not filePath:
            return

        try:
            with open(filePath, 'r', encoding='utf-8') as file:
                lines = file.readlines()

                if not lines:
                    QMessageBox.warning(self, "Warning", "The file is empty!")
                    return

                wb = Workbook()

                for i, line in enumerate(lines, start=1):
                    sheet = wb.create_sheet(title=f"Sheet{i}")
                    sheet.append([line.strip()])

                output_path = os.path.splitext(filePath)[0] + ".xlsx"
                wb.remove(wb[wb.sheetnames[0]])  # Remove default sheet
                wb.save(output_path)

                QMessageBox.information(self, "Success", f"File converted successfully!\nSaved as: {output_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to process file: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    converter = TxtToExcelConverter()
    converter.show()
    sys.exit(app.exec_())
